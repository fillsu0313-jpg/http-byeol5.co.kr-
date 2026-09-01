"""
7월 엑셀 데이터 → DB 시딩 (검증/시연용)
API 연동 완료 후 폐기 가능.

실행: python scripts/seed_july.py
필요 파일: files/7월_상품별_흐름__1.xlsx
"""
import hashlib
import sqlite3
import sys
from datetime import date, datetime
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))

DB_PATH = BASE_DIR / "data" / "coupang.db"
EXCEL_PATH = BASE_DIR / "files" / "7월_상품별_흐름__1.xlsx"


def make_vendor_item_id(name: str) -> int:
    """상품명 → 임시 vendor_item_id (init_db.py와 동일 해시)"""
    h = int(hashlib.sha256(name.encode("utf-8")).hexdigest(), 16)
    return h % 10_000_000 + 100_000


def parse_date(raw) -> str:
    """다양한 날짜 형식 → '2026-07-15'"""
    # datetime 객체
    if isinstance(raw, (date, datetime)):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    # '2026-07-15' 형식 그대로
    if len(s) == 10 and "-" in s:
        return s
    # 점(.) 구분 형식: '7.15', '2026.7.1', '2026.07.15' 등
    if "." in s:
        parts = s.split(".")
        if len(parts) == 2:
            # '7.15' → 월.일
            month = int(parts[0])
            day = int(parts[1])
            return f"2026-{month:02d}-{day:02d}"
        elif len(parts) == 3:
            # '2026.7.1' → 년.월.일
            year = int(parts[0])
            month = int(parts[1])
            day = int(parts[2])
            return f"{year}-{month:02d}-{day:02d}"
    raise ValueError(f"날짜 변환 실패: {raw}")


def seed():
    if not EXCEL_PATH.exists():
        print(f"엑셀 파일이 없습니다: {EXCEL_PATH}")
        print("files/ 폴더에 '7월_상품별_흐름__1.xlsx'를 배치하세요.")
        sys.exit(1)

    if not DB_PATH.exists():
        print(f"DB가 없습니다: {DB_PATH}")
        print("먼저 python scripts/init_db.py 를 실행하세요.")
        sys.exit(1)

    try:
        import openpyxl
    except ImportError:
        print("openpyxl이 필요합니다: pip install openpyxl")
        sys.exit(1)

    # 엑셀 로드
    wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
    ws = wb["상품별흐름-판매"]

    # 헤더 찾기
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(c or "").strip() for c in row]
        break

    col = {h: i for i, h in enumerate(headers)}
    required = ["상품명", "일자"]
    for r in required:
        if r not in col:
            print(f"필수 컬럼 '{r}'이(가) 없습니다. 헤더: {headers}")
            sys.exit(1)

    # DB에 등록된 상품명 → vendor_item_id 매핑 로드
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")

    products = conn.execute("SELECT vendor_item_id, display_name FROM products").fetchall()
    name_to_vid = {r["display_name"]: r["vendor_item_id"] for r in products}

    # 해시 기반 매핑도 준비 (init_db.py와 동일 방식)
    # 충돌 해결 포함
    id_to_name = {r["vendor_item_id"]: r["display_name"] for r in products}

    sales_count = 0
    ads_count = 0
    skip_count = 0
    warnings = []
    unknown_products = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = list(row)
        name = cells[col["상품명"]]
        if not name or str(name).strip() == "":
            continue
        name = str(name).strip()

        # 상품명 → vendor_item_id
        vid = name_to_vid.get(name)
        if vid is None:
            # 해시로 시도
            vid_hash = make_vendor_item_id(name)
            if vid_hash in id_to_name:
                vid = vid_hash
            else:
                if name not in unknown_products:
                    unknown_products.add(name)
                    warnings.append(f"미등록 상품: '{name}'")
                skip_count += 1
                continue

        # 날짜 파싱
        raw_date = cells[col["일자"]]
        if raw_date is None:
            skip_count += 1
            continue
        try:
            stat_date = parse_date(raw_date)
        except ValueError as e:
            warnings.append(str(e))
            skip_count += 1
            continue

        # 숫자 컬럼 읽기
        def get_num(col_name):
            idx = col.get(col_name)
            if idx is None or idx >= len(cells):
                return None
            v = cells[idx]
            if v is None or str(v).strip() == "":
                return None
            try:
                return float(v)
            except (ValueError, TypeError):
                return None

        units_sold = get_num("판매량")
        ad_units = get_num("광고")
        total_views = get_num("총 조회수")
        ad_clicks = get_num("광고 클릭수")
        ad_spend_raw = get_num("광고비")  # 엑셀의 광고비는 ×1.1 포함

        # daily_sales 삽입
        conn.execute(
            """INSERT OR REPLACE INTO daily_sales
               (stat_date, vendor_item_id, units_sold, total_views, source)
               VALUES (?, ?, ?, ?, 'seed')""",
            (
                stat_date,
                vid,
                int(units_sold) if units_sold is not None else None,
                int(total_views) if total_views is not None else None,
            ),
        )
        sales_count += 1

        # daily_ads 삽입 (광고비 ÷ 1.1로 원본값 복원)
        ad_spend = None
        if ad_spend_raw is not None:
            ad_spend = ad_spend_raw / 1.1  # 원본값 복원

        conn.execute(
            """INSERT OR REPLACE INTO daily_ads
               (stat_date, vendor_item_id, ad_units, ad_clicks, ad_spend, source)
               VALUES (?, ?, ?, ?, ?, 'seed')""",
            (
                stat_date,
                vid,
                int(ad_units) if ad_units is not None else None,
                int(ad_clicks) if ad_clicks is not None else None,
                round(ad_spend, 2) if ad_spend is not None else None,
            ),
        )
        ads_count += 1

    conn.commit()

    # ingest_log 기록
    conn.execute(
        """INSERT INTO ingest_log
           (stat_date, data_type, source, row_count, status, message)
           VALUES ('2026-07', 'sales', 'seed', ?, 'ok', ?)""",
        (sales_count, f"7월 엑셀 시딩"),
    )
    conn.execute(
        """INSERT INTO ingest_log
           (stat_date, data_type, source, row_count, status, message)
           VALUES ('2026-07', 'ads', 'seed', ?, 'ok', ?)""",
        (ads_count, f"7월 엑셀 시딩"),
    )
    conn.commit()
    conn.close()
    wb.close()

    print(f"=== 7월 시딩 완료 ===")
    print(f"  daily_sales: {sales_count}행 삽입")
    print(f"  daily_ads:   {ads_count}행 삽입")
    print(f"  건너뜀:      {skip_count}행")
    if warnings:
        print(f"\n=== 경고 ({len(warnings)}건) ===")
        for w in warnings[:20]:
            print(f"  {w}")
        if len(warnings) > 20:
            print(f"  ... 외 {len(warnings) - 20}건")


if __name__ == "__main__":
    seed()
